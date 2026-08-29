from __future__ import annotations

import io
import re
import requests
import urllib3
from openpyxl import load_workbook
import pymupdf

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

URLS = {
    "housing_ch3": "https://www.knbs.or.ke/wp-content/uploads/2025/04/Chapter-3-Household-Demographic-and-Economic-Characteristics.xlsx",
    "housing_ch5": "https://www.knbs.or.ke/wp-content/uploads/2025/04/Chapter-5-Housing-Characteristics-Amenities-and-Adequacy.xlsx",
    "gcp": "https://www.knbs.or.ke/wp-content/uploads/2025/12/2025-Gross-County-Product.pdf",
    "agri": "https://www.knbs.or.ke/wp-content/uploads/2025/01/National-Agriculture-Production-Report-2024.pdf",
}


def get(url: str) -> bytes:
    # KNBS currently serves an incomplete public TLS chain to some Linux runners.
    # This diagnostic reads only the explicitly pinned official knbs.or.ke URLs above.
    r = requests.get(url, timeout=120, verify=False, headers={"User-Agent": "Kenya-Data-Atlas/0.10 source-validation"})
    r.raise_for_status()
    print(f"DOWNLOAD_OK {url} bytes={len(r.content)} content_type={r.headers.get('content-type')}")
    return r.content


def compact_row(row):
    vals = []
    for i, v in enumerate(row, 1):
        if v is not None and str(v).strip() != "":
            vals.append(f"C{i}={v}")
    return " | ".join(vals)


def target_xlsx(label: str, content: bytes, targets: list[str]):
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    print(f"XLSX {label} sheets={len(wb.sheetnames)}")
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        hits = []
        for idx, row in enumerate(rows):
            joined = " ".join("" if v is None else str(v) for v in row)
            if any(t.lower() in joined.lower() for t in targets) or any(t.lower() in ws.title.lower() for t in targets):
                hits.append(idx)
        if not hits:
            continue
        print(f"TARGET_SHEET {label} title={ws.title!r} rows={ws.max_row} cols={ws.max_column} hit_rows={[x+1 for x in hits[:20]]}")
        lo = max(0, min(hits) - 6)
        hi = min(len(rows), max(hits) + 60)
        for j in range(lo, hi):
            line = compact_row(rows[j])
            if line:
                print(f"  R{j+1}: {line[:3000]}")


def show_pdf(label: str, content: bytes, needles: list[str]):
    doc = pymupdf.open(stream=content, filetype="pdf")
    print(f"PDF {label} pages={doc.page_count}")
    for pno in range(doc.page_count):
        text = doc[pno].get_text("text")
        compact = re.sub(r"\s+", " ", text)
        if any(n.lower() in compact.lower() for n in needles):
            print(f"MATCH {label} page={pno+1}")
            print(text[:9000])


def main():
    target_xlsx("housing_ch3", get(URLS["housing_ch3"]), ["Table 3.18", "Table 3.19", "Used Internet", "Used a Computer"])
    target_xlsx("housing_ch5", get(URLS["housing_ch5"]), ["Table 5.11", "Connected to Electricity", "Main Grid", "electricity"])
    # PDFs are already structurally confirmed; keep concise page checks so URLs remain exercised.
    show_pdf("gcp", get(URLS["gcp"]), ["GCP by Economic Activity at Current Prices, 2024"])
    show_pdf("agri", get(URLS["agri"]), ["Area and Production of Maize by County"])
    print("P05_SOURCE_DIAGNOSTIC_OK")


if __name__ == "__main__":
    main()
