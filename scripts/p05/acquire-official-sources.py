from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path

import pymupdf
import requests
import urllib3
from openpyxl import load_workbook

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / 'data' / 'p05' / 'source'
OUT.mkdir(parents=True, exist_ok=True)

URLS = {
    'housing_ch3': 'https://www.knbs.or.ke/wp-content/uploads/2025/04/Chapter-3-Household-Demographic-and-Economic-Characteristics.xlsx',
    'housing_ch5': 'https://www.knbs.or.ke/wp-content/uploads/2025/04/Chapter-5-Housing-Characteristics-Amenities-and-Adequacy.xlsx',
    'gcp': 'https://www.knbs.or.ke/wp-content/uploads/2025/12/2025-Gross-County-Product.pdf',
    'agri': 'https://www.knbs.or.ke/wp-content/uploads/2025/01/National-Agriculture-Production-Report-2024.pdf',
}


def norm(s):
    s = str(s or '').upper().replace('’', "'").replace('–', '-').replace('\n', ' ')
    return re.sub(r'[^A-Z0-9]+', '', s)


def canonical_counties():
    rows = []
    with open(ROOT/'data/geography/registry/geographies.csv', encoding='utf-8') as f:
        for r in csv.DictReader(f):
            if r.get('level') == 'county':
                rows.append((r['geo_code'], r['name']))
    if len(rows) != 47:
        raise RuntimeError(f'Expected 47 registry counties, got {len(rows)}')
    return rows

COUNTIES = canonical_counties()
NAME_TO_GEO = {norm(name): (geo, name) for geo, name in COUNTIES}
for alias, target in {'HOMABAY':'Homa Bay','NAIROBI':'Nairobi City'}.items():
    tgt = next((x for x in COUNTIES if x[1] == target), None)
    if tgt:
        NAME_TO_GEO[alias] = tgt


def fetch(url):
    # KNBS public servers currently present an incomplete TLS chain to some Linux runners.
    # Acquisition is restricted to the pinned official HTTPS URLs above.
    r = requests.get(url, timeout=120, verify=False, headers={'User-Agent':'Kenya-Data-Atlas/0.10 P05-acquisition'})
    r.raise_for_status()
    return r.content


def number(v):
    if v is None or str(v).strip() in {'', '-', '—'}:
        return None
    s = str(v).strip().replace(',', '')
    if s.startswith('(') and s.endswith(')'):
        s = '-' + s[1:-1]
    return float(s)


def put_county(out, source_name, raw_value):
    key = norm(source_name)
    if key not in NAME_TO_GEO:
        return False
    geo, name = NAME_TO_GEO[key]
    value = number(raw_value)
    if value is None:
        raise RuntimeError(f'Missing numeric value for {name}')
    if geo in out and out[geo] != value:
        raise RuntimeError(f'Conflicting duplicate for {name}: {out[geo]} vs {value}')
    out[geo] = value
    return True


def parse_two_panel_sheet(ws, value_left_col, value_right_col):
    out = {}
    for row in ws.iter_rows(values_only=True):
        if len(row) >= value_left_col:
            put_county(out, row[0], row[value_left_col-1])
        if len(row) >= value_right_col:
            put_county(out, row[5], row[value_right_col-1])
    if len(out) != 47:
        missing = [n for g,n in COUNTIES if g not in out]
        raise RuntimeError(f'{ws.title}: expected 47 counties, got {len(out)} missing={missing}')
    return out


def parse_connectivity(ch3_bytes, ch5_bytes):
    wb3 = load_workbook(io.BytesIO(ch3_bytes), data_only=True, read_only=True)
    internet = parse_two_panel_sheet(wb3['Table 3.18'], 4, 9)
    computer = parse_two_panel_sheet(wb3['Table 3.19'], 4, 9)
    wb5 = load_workbook(io.BytesIO(ch5_bytes), data_only=True, read_only=True)
    electricity = {}
    for row in wb5['Table 5.11'].iter_rows(values_only=True):
        if len(row) >= 2:
            put_county(electricity, row[0], row[1])
    if len(electricity) != 47:
        raise RuntimeError(f'Electricity expected 47 rows, got {len(electricity)}')
    return internet, computer, electricity


def parse_gcp(pdf_bytes):
    doc = pymupdf.open(stream=pdf_bytes, filetype='pdf')
    out = {}
    for page in doc:
        text = page.get_text('text')
        if 'GCP' not in text:
            continue
        try:
            tables = page.find_tables().tables
        except Exception:
            continue
        for table in tables:
            for row in table.extract():
                if len(row) < 22:
                    continue
                key = norm(row[1])
                if key not in NAME_TO_GEO:
                    continue
                geo, name = NAME_TO_GEO[key]
                rec = {
                    'agriculture_gva_ksh_m': number(row[2]),
                    'manufacturing_gva_ksh_m': number(row[4]),
                    'gcp_ksh_m': number(row[21]),
                }
                if any(v is None for v in rec.values()):
                    raise RuntimeError(f'Incomplete GCP row {name}: {row}')
                rec['agriculture_share_pct'] = round(rec['agriculture_gva_ksh_m']/rec['gcp_ksh_m']*100, 2)
                rec['manufacturing_share_pct'] = round(rec['manufacturing_gva_ksh_m']/rec['gcp_ksh_m']*100, 2)
                out[geo] = rec
    if len(out) != 47:
        missing = [n for g,n in COUNTIES if g not in out]
        raise RuntimeError(f'GCP expected 47, got {len(out)} missing={missing}')
    existing = {}
    with open(ROOT/'data/sprint1/gcp-2020-2024.csv', encoding='utf-8') as f:
        for r in csv.DictReader(f):
            existing[r['geo_code']] = float(r['2024'])
    mismatches = [(g,out[g]['gcp_ksh_m'],existing.get(g)) for g,_ in COUNTIES if out[g]['gcp_ksh_m'] != existing.get(g)]
    if mismatches:
        raise RuntimeError(f'GCP 2024 reconciliation mismatch: {mismatches[:5]}')
    return out


def parse_maize(pdf_bytes):
    doc = pymupdf.open(stream=pdf_bytes, filetype='pdf')
    out = {}
    for page in doc:
        text = page.get_text('text')
        if 'Area and Production of Maize by County' not in text:
            continue
        for table in page.find_tables().tables:
            names = list(table.header.names or [])
            rows = table.extract()
            area = next((r for r in rows if len(r)>1 and str(r[0]).strip()=='2023' and 'Area' in str(r[1])), None)
            prod = None
            if area is not None:
                ai = rows.index(area)
                if ai > 0 and 'Production' in str(rows[ai-1][1]):
                    prod = rows[ai-1]
            if area is None or prod is None:
                continue
            for i in range(2, min(len(names), len(area), len(prod))):
                key = norm(names[i])
                if key not in NAME_TO_GEO:
                    continue
                geo, _ = NAME_TO_GEO[key]
                a, p = number(area[i]), number(prod[i])
                if a is None or p is None:
                    continue
                out[geo] = {'maize_area_ha':a,'maize_production_tonnes':p,'maize_yield_t_per_ha':round(p/a,3) if a else None}
    if len(out) != 47:
        print('AGRI_HEADER_DIAGNOSTIC')
        for page in doc:
            if 'Area and Production of Maize by County' in page.get_text('text'):
                for table in page.find_tables().tables:
                    print(table.header.names)
        missing = [n for g,n in COUNTIES if g not in out]
        raise RuntimeError(f'Maize expected 47, got {len(out)} missing={missing}')
    total_area = round(sum(r['maize_area_ha'] for r in out.values()))
    total_prod = round(sum(r['maize_production_tonnes'] for r in out.values()))
    if total_area != 2430013 or total_prod != 4285206:
        raise RuntimeError(f'Maize 2023 reconciliation failed area={total_area} prod={total_prod}')
    return out


def parse_education():
    out = {}
    with open(ROOT/'data/place-facts/source/county-key-facts.csv', encoding='utf-8') as f:
        for r in csv.DictReader(f):
            out[r['geo_code']] = {
                'public_primary_schools':int(r['public_primary_schools']),
                'primary_classroom_teachers':int(r['primary_classroom_teachers']),
                'public_secondary_schools':int(r['public_secondary_schools']),
                'secondary_teachers':int(r['secondary_teachers']),
            }
    if len(out) != 47:
        raise RuntimeError('Education source must contain 47 counties')
    totals = {k:sum(x[k] for x in out.values()) for k in next(iter(out.values()))}
    expected = {'public_primary_schools':23274,'primary_classroom_teachers':183929,'public_secondary_schools':9246,'secondary_teachers':108569}
    if totals != expected:
        raise RuntimeError(f'Education reconciliation failed {totals}')
    return out


def write_csv(path, fields, by_geo):
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=['county_number','geo_code','name']+fields)
        w.writeheader()
        for idx,(geo,name) in enumerate(COUNTIES,1):
            w.writerow({'county_number':f'{idx:03d}','geo_code':geo,'name':name,**by_geo[geo]})


def main():
    ch3, ch5 = fetch(URLS['housing_ch3']), fetch(URLS['housing_ch5'])
    internet, computer, electricity = parse_connectivity(ch3,ch5)
    connectivity = {g:{'internet_use_pct':internet[g],'computer_use_pct':computer[g],'main_grid_electricity_pct':electricity[g]} for g,_ in COUNTIES}
    gcp = parse_gcp(fetch(URLS['gcp']))
    maize = parse_maize(fetch(URLS['agri']))
    education = parse_education()

    write_csv(OUT/'education-tsc-2023.csv', list(next(iter(education.values())).keys()), education)
    write_csv(OUT/'economic-structure-gcp-2024.csv', list(next(iter(gcp.values())).keys()), gcp)
    write_csv(OUT/'maize-2023.csv', list(next(iter(maize.values())).keys()), maize)
    write_csv(OUT/'connectivity-housing-survey-2023-24.csv', list(next(iter(connectivity.values())).keys()), connectivity)
    manifest = {'sources':URLS,'county_count':47,'reconciliations':{
        'education':{'public_primary_schools':23274,'primary_classroom_teachers':183929,'public_secondary_schools':9246,'secondary_teachers':108569},
        'maize_2023':{'area_ha':2430013,'production_tonnes':4285206},
        'gcp_2024':'reconciled exactly to data/sprint1/gcp-2020-2024.csv',
        'connectivity':'47 county rows per metric'
    }}
    (OUT/'manifest.json').write_text(json.dumps(manifest,indent=2)+'\n',encoding='utf-8')
    print('P05_ACQUISITION_47_COUNTIES_OK')
    print('P05_EDUCATION_RECONCILIATION_OK')
    print('P05_GCP_RECONCILIATION_OK')
    print('P05_MAIZE_RECONCILIATION_OK area=2430013 production_tonnes=4285206')
    print('P05_CONNECTIVITY_RECONCILIATION_OK')

if __name__=='__main__':
    main()
